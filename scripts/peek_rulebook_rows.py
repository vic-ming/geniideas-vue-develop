from openpyxl import load_workbook


def norm(v):
    return "" if v is None else str(v).strip()


def peek(sheet_name, rows, max_col=26):
    path = "src/assets/export/氣體別、聚賢料表、料號編碼規則 20251202.xlsx"
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    for r in rows:
        values = [norm(ws.cell(r, c).value) for c in range(1, max_col + 1)]
        print(f"{sheet_name} row {r}:")
        print(list(enumerate(values, start=1)))
        print()


def main():
    peek("配件 (聚賢)", rows=[2, 2210, 2211, 2212, 2213, 2214, 4616, 4617, 4810, 4811, 4815, 4816], max_col=28)
    peek("其他元件 (聚賢)", rows=[1, 2, 3, 9, 10, 14, 15, 20, 21, 26, 112, 113, 404, 405, 696, 697], max_col=32)


if __name__ == "__main__":
    main()


