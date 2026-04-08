import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path("src/assets/export")
RULE_BOOK = ROOT / "氣體別、聚賢料表、料號編碼規則 20251202.xlsx"
OUT = ROOT / "juxian_pipe_rules.json"

wb = load_workbook(RULE_BOOK, read_only=True, data_only=True)
ws = wb["管線 (聚賢)"]


def norm(v):
    return "" if v is None else str(v).strip()


# header row defines brand pairs starting from column H (0-based 7)
header = [norm(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
brand_cols = []  # (brand, name_col_idx, part_col_idx)
for idx in range(7, len(header) - 1, 2):
    b1 = header[idx]
    b2 = header[idx + 1]
    if b1 and b1 == b2:
        brand_cols.append((b1, idx, idx + 1))

rules = []
unmatched = []

for row in ws.iter_rows(min_row=2, values_only=True):
    r = list(row)
    if len(r) < 7:
        continue
    pipeline_type = norm(r[2])
    size = norm(r[4])
    material = norm(r[6])

    if not pipeline_type or not size or not material:
        continue

    for brand, n_idx, p_idx in brand_cols:
        name = norm(r[n_idx]) if n_idx < len(r) else ""
        part_no = norm(r[p_idx]) if p_idx < len(r) else ""
        if not name and not part_no:
            continue
        entry = {
            "pipelineType": pipeline_type,
            "size": size,
            "material": material,
            "brand": brand,
            "name": name,
            "partNo": part_no,
            "unit": "M",
        }
        if name == "無可對應料號" or part_no == "無可對應料號" or not name or not part_no:
            unmatched.append(entry)
        else:
            rules.append(entry)

OUT.write_text(json.dumps({"rules": rules, "unmatched": unmatched}, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {len(rules)} rules, {len(unmatched)} unmatched -> {OUT}")
