from openpyxl import load_workbook


def norm(v):
    return "" if v is None else str(v).strip()


def scan_sheet(ws, max_row=260, max_col=30):
    print(f"sheet={ws.title} max_col={ws.max_column} max_row={ws.max_row}")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        r = [norm(x) for x in row]
        if not any(r):
            continue
        rep = 0
        for j in range(len(r) - 1):
            if r[j] and r[j] == r[j + 1]:
                rep += 1
        title = r[1] if len(r) > 1 else ""
        maybe_title = bool(title and any(k in title for k in ("REDUCER", "TEE", "ELBOW", "DOUBLE", "GAUGE", "SPRING", "GLAND", "NUT", "GASKET", "STOP")))
        maybe_size = bool(len(r) > 2 and r[2] and '"' in r[2])
        maybe_header = rep >= 2
        if maybe_title or maybe_size or maybe_header:
            print(i, "rep", rep, "|", r[:max_col])


def main():
    path = "src/assets/export/氣體別、聚賢料表、料號編碼規則 20251202.xlsx"
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in ["配件 (聚賢)", "其他元件 (聚賢)"]:
        print("\n====", sheet, "====")
        scan_sheet(wb[sheet], max_row=1600, max_col=28)


if __name__ == "__main__":
    main()


