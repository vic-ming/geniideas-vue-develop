import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path("src/assets/export")
RULE_BOOK = ROOT / "氣體別、聚賢料表、料號編碼規則 20251202.xlsx"

def norm(v):
    return "" if v is None else str(v).strip()

def write_json(path: Path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {path.name}")

# --- TSMC Extractors ---

def extract_tsmc_pipe(wb):
    ws = wb["管線"]
    rules, unmatched = [], []
    last_category = last_double_size = last_panel_size = ""
    for row in ws.iter_rows(values_only=True):
        r = [norm(x) for x in row]
        if len(r) < 10: r += [""] * (10 - len(r))
        
        c_sec, c_cat, c_ds, c_ps = r[1], r[2], r[3], r[4]
        if c_sec and not c_cat: last_category = c_sec; continue
        if c_cat: last_category = c_cat
        if c_ds: last_double_size = c_ds
        if c_ps: last_panel_size = c_ps
        
        cat = last_category or c_cat
        ds = last_double_size if cat == "雙套管" else ""
        ps = c_ps if c_ps else (last_panel_size if cat != "雙套管" else "")
        mat, part, length_hint, qty_hint, note = r[6], r[7], r[5], r[8], r[9]
        
        if not part or part == "台積料表": continue
        
        entry = {
            "match": { "pipelineType": [cat], "material": [mat] },
            "output": { "partName": part, "lengthHint": length_hint, "qtyHint": qty_hint, "note": note, "unit": "M" }
        }
        if ds: entry["match"]["doubleSize"] = [ds]
        if ps: entry["match"]["panelSize"] = [ps]
        
        if part == "無可對應料號":
            unmatched.append(entry["match"])
        else:
            rules.append(entry)
    return {"rules": rules, "unmatched": unmatched}

def extract_tsmc_valve(wb):
    ws = wb["閥件"]
    rules, unmatched = [], []
    for row in ws.iter_rows(values_only=True):
        r = [norm(x) for x in row]
        if len(r) < 9: continue
        enabled, conn, size, vtype, part, qty = r[2], r[3], r[4], r[5], r[7], r[8]
        if enabled != "勾選" or not conn or not size or not vtype: continue
        if not part or part == "台積料表": continue
        
        entry = {
            "match": { "connector": [conn], "size": [size], "valveType": [vtype] },
            "output": { "partName": part, "unit": "EA", "qtyHint": qty }
        }
        if part == "無可對應料號": unmatched.append(entry["match"])
        else: rules.append(entry)
    return {"rules": rules, "unmatched": unmatched}

def extract_tsmc_fitting(wb):
    ws = wb["配件"]
    rules, unmatched = [], []
    for row in ws.iter_rows(values_only=True):
        r = [norm(x) for x in row]
        if len(r) < 10: r += [""] * (10 - len(r))
        size, size2, mat, part, qty_hint = r[2], r[3], r[4], r[7], r[9]
        if not part or part in ("台積料表", "料表", "配件品名"): continue
        
        m_type = ""
        m_size = size
        d_size = ""
        
        pu = part.upper()
        if "90 DEG" in pu or "90 DI" in pu: m_type = "90 DI ELBOW"
        elif "45 DEG" in pu or "45 DI" in pu: m_type = "45 DI ELBOW"
        elif "REDUCER TEE" in pu or "REDUCING TEE" in pu: m_type = "REDUCING TEE"; m_size = f"{size}x{size2}"
        elif "REDUCER" in pu: m_type = "REDUCER"; m_size = f"{size}x{size2}"
        elif "EQUAL TEE" in pu or ("TEE 有分支" in r[1]): m_type = "TEE"
        elif "DOUBLE TEE" in pu: m_type = "TEE"; d_size = size; m_size = ""
        elif "DOUBLE ELBOW" in pu: m_type = "90 DI ELBOW"; d_size = size; m_size = ""
        elif "CROSS" in pu: m_type = "CROSS"
        elif "CAP" in pu: m_type = "CAP"
        
        entry = {
            "match": {
                "pipelineType": ["單套管", "單線管", "盤面"] if not d_size else ["雙套管"],
                "fittingType": [m_type],
                "connector": ["WELD"],
                "material": [mat]
            },
            "output": { "partName": part, "unit": "EA" }
        }
        if m_size: entry["match"]["size"] = [m_size]
        if d_size: entry["match"]["doubleSize"] = [d_size]
        
        # ELBOW divisor rule
        if "ELBOW" in pu and not d_size:
            entry["output"]["divisor"] = 3
            entry["output"]["round"] = "ceil" if "進位" in qty_hint else "floor"
            
        if part == "無可對應料號": unmatched.append(entry["match"])
        else: rules.append(entry)
    return {"rules": rules, "unmatched": unmatched}

def extract_tsmc_other(wb):
    ws = wb["其他元件"]
    stopSpacer, spring, overTube, gauge, gland, nut, gasket = [], [], [], [], [], [], []
    current = None
    for row in ws.iter_rows(values_only=True):
        r = [norm(x) for x in row]
        if len(r) < 10: r += [""] * (10 - len(r))
        label = r[1]
        if label.startswith("STOP SPACER"): current = stopSpacer; continue
        if label.startswith("SPRING"): current = spring; continue
        if label.startswith("Over Tube"): current = overTube; continue
        if label.startswith("GAUGE"): current = gauge; continue
        if label == "GLAND": current = gland; continue
        if label == "NUT": current = nut; continue
        if label == "GASKET": current = gasket; continue
        
        if current is stopSpacer or current is spring or current is overTube:
            ptype, ds, part = r[2], r[3], r[7]
            if ptype == "雙套管" and ds and part and part != "無可對應料號":
                current.append({ "match": { "doubleSize": [ds] }, "output": { "partName": part, "unit": "EA" } })
        elif current is gauge or current is gland or current is nut or current is gasket:
            ps, pc, mat, part = r[3], r[4], r[5], r[7]
            if ps and pc and mat and part and part != "無可對應料號":
                current.append({ "match": { "panelSize": [ps], "panelConnector": [pc], "material": [mat] }, "output": { "partName": part, "unit": "EA" } })
                
    return { "stopSpacer": stopSpacer, "spring": spring, "overTube": overTube, "gauge": gauge, "gland": gland, "nut": nut, "gasket": gasket }

# --- Juxian Extractors ---

def scan_adjacent_pairs(row_values, start_idx):
    for i in range(start_idx, len(row_values) - 1):
        name = row_values[i]
        part_no = row_values[i + 1]
        if not name or not part_no: continue
        if "氣體材料" not in name and "G" not in part_no: continue
        yield name, part_no

def guess_brand(name):
    parts = name.split("_")
    return parts[-2].strip() if len(parts) >= 2 else ""

def extract_juxian_pipe(wb):
    ws = wb["管線 (聚賢)"]
    rules, unmatched = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = [norm(x) for x in row]
        if len(r) < 7: continue
        ptype, size, mat = r[2], r[4], r[6]
        if not ptype or not size or not mat: continue
        
        for name, part_no in scan_adjacent_pairs(r, start_idx=7):
            entry = {
                "match": { "pipelineType": [ptype], "size": [size], "material": [mat], "brand": [guess_brand(name)] },
                "output": { "name": name, "partNo": part_no, "unit": "M" }
            }
            if name == "無可對應料號" or part_no == "無可對應料號": unmatched.append(entry["match"])
            else: rules.append(entry)
    return {"rules": rules, "unmatched": unmatched}

def extract_juxian_valve(wb):
    ws = wb["閥件 (聚賢)"]
    rules, unmatched = [], []
    header = [norm(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    brand_cols = []
    for idx in range(len(header) - 1):
        if header[idx] and header[idx] == header[idx+1]: brand_cols.append((header[idx], idx, idx+1))
        
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = [norm(x) for x in row]
        if len(r) < 6: continue
        checked, conn, size, vtype_raw = r[2], r[3], r[4], r[5]
        if checked != "勾選" or not conn or not size: continue
        
        for brand, n_idx, p_idx in brand_cols:
            if n_idx >= len(r) or p_idx >= len(r): continue
            name, part_no = r[n_idx], r[p_idx]
            if not name and not part_no: continue
            
            vtype = vtype_raw
            if "REGULATOR" in name.upper(): vtype = "REGULATOR"
            if not vtype or vtype == "勾選": continue
            
            entry = {
                "match": { "connector": [conn], "size": [size], "valveType": [vtype], "brand": [brand] },
                "output": { "name": name, "partNo": part_no, "unit": "EA" }
            }
            if name == "無可對應料號" or part_no == "無可對應料號": unmatched.append(entry["match"])
            else: rules.append(entry)
    return {"rules": rules, "unmatched": unmatched}

def extract_juxian_fitting(wb):
    ws = wb["配件 (聚賢)"]
    rules, unmatched = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = [norm(x) for x in row]
        if len(r) < 6: continue
        
        # 聚賢配件 usually has label in r[1], but sizes are scattered.
        # However, the user places REDUCER in F, G, H, I, J!
        # WHICH columns does Juxian use?
        # A=1, B=2, C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10
        # If it's a completely flat structure across the sheet: Let's extract EVERY pair of columns that look like name & part!
        
        title = r[1]
        base_match = {}
        if "REDUCER 無分支" in title: base_match = { "type": ["reducer"], "fromSize": [r[2]], "toSize": [r[3]], "material": [r[4]] }
        elif "REDUCER TEE" in title: base_match = { "type": ["reducerTee"], "fromSize": [r[2]], "toSize": [r[3]], "material": [r[4]] }
        elif "TEE 有分支" in title: base_match = { "type": ["tee"], "size": [r[3]], "material": [r[4]] }
        elif "ELBOW" in title: base_match = { "type": ["elbow"], "size": [r[2]], "material": [r[4]] }
        elif "DOUBLE TEE" in title: base_match = { "type": ["doubleTee"], "pipelineType": [r[2]], "doubleSize": [r[3]] }
        elif "DOUBLE ELBOW" in title: base_match = { "type": ["doubleElbow"], "pipelineType": [r[2]], "doubleSize": [r[3]] }
        # Reducers at columns F, G, H, I, J ?
        # We can detect dynamically:
        title2 = r[5]
        base_match2 = {}
        if "REDUCER TEE" in title2: base_match2 = { "type": ["reducerTee"], "fromSize": [r[6]], "toSize": [r[7]], "material": [r[8]] }
        elif "REDUCER" in title2: base_match2 = { "type": ["reducer"], "fromSize": [r[6]], "toSize": [r[7]], "material": [r[8]] }
        elif "TEE" in title2: base_match2 = { "type": ["tee"], "size": [r[7]], "material": [r[8]] }
        elif "ELBOW" in title2: base_match2 = { "type": ["elbow"], "size": [r[6]], "material": [r[8]] }
        
        if base_match:
            for n, p in scan_adjacent_pairs(r, 6):
                rules.append({"match": {**base_match, "brand": [guess_brand(n)]}, "output": {"name": n, "partNo": p, "unit": "EA"}})
        if base_match2:
            for n, p in scan_adjacent_pairs(r, 9):
                rules.append({"match": {**base_match2, "brand": [guess_brand(n)]}, "output": {"name": n, "partNo": p, "unit": "EA"}})
                
    return {"rules": rules, "unmatched": unmatched}

def extract_juxian_other(wb):
    ws = wb["其他元件 (聚賢)"]
    rules, unmatched = [], []
    current = "stopSpacer"
    for row in ws.iter_rows(min_row=1, values_only=True):
        r = [norm(x) for x in row]
        if not any(r): continue
        t = r[1]
        if "STOP SPACER" in t: current = "stopSpacer"; continue
        if "SPRING" in t: current = "spring"; continue
        if "Over Tube" in t: current = "overTube"; continue
        if "GAUGE" in t: current = "gauge"; continue
        if "GLAND" in t: current = "gland"; continue
        if "NUT" in t: current = "nut"; continue
        if "GASKET" in t: current = "gasket"; continue
        if len(r) > 2 and r[2].startswith("A2"): continue
        
        base = {"section": [current]}
        if current in ("stopSpacer", "spring", "overTube"):
            base["pipelineType"], base["doubleSize"] = [r[2]], [r[3]]
            if current == "stopSpacer" and len(r) > 5 and r[5] == "勾選": base["hasBranchPanel"] = [True]
        elif current in ("gauge", "gland", "nut", "gasket"):
            base["panelSize"], base["panelConnector"], base["material"] = [r[3]], [r[4]], [r[5]]
        else: continue
        
        if not base.get("panelSize") and not base.get("doubleSize"): continue # Filter empty
        
        for n, p in scan_adjacent_pairs(r, 7):
            rules.append({"match": {**base, "brand": [guess_brand(n)]}, "output": {"name": n, "partNo": p, "unit": "EA"}})
    return {"rules": rules, "unmatched": unmatched}

def main():
    wb = load_workbook(RULE_BOOK, read_only=True, data_only=True)
    write_json(ROOT / "tsmc_pipe_rules.json", extract_tsmc_pipe(wb))
    write_json(ROOT / "tsmc_valve_rules.json", extract_tsmc_valve(wb))
    write_json(ROOT / "tsmc_fitting_rules.json", extract_tsmc_fitting(wb))
    write_json(ROOT / "tsmc_other_rules.json", extract_tsmc_other(wb))
    write_json(ROOT / "juxian_pipe_rules.json", extract_juxian_pipe(wb))
    write_json(ROOT / "juxian_valve_rules.json", extract_juxian_valve(wb))
    write_json(ROOT / "juxian_fitting_rules.json", extract_juxian_fitting(wb))
    write_json(ROOT / "juxian_other_rules.json", extract_juxian_other(wb))

if __name__ == "__main__":
    main()
