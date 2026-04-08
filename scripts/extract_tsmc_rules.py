import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path('src/assets/export')
RULE_BOOK = ROOT / '氣體別、聚賢料表、料號編碼規則 20251202.xlsx'
OUTPUT = ROOT / 'tsmc_pipe_rules.json'

wb = load_workbook(RULE_BOOK, read_only=True, data_only=True)
ws = wb['管線']

pipe_rules = []
unmatched = []

last_category = None
last_double_size = None
last_panel_size = None

for row in ws.iter_rows(values_only=True):
    values = list(row)
    if len(values) < 10:
        values += [None] * (10 - len(values))

    raw_section = (values[1] or '').strip()
    raw_category = (values[2] or '').strip()
    raw_double_size = (values[3] or '').strip()
    raw_panel_size = (values[4] or '').strip()
    length_hint = (values[5] or '').strip()
    raw_material = (values[6] or '').strip()
    part_name = (values[7] or '').strip()
    qty_hint = (values[8] or '').strip()
    note = (values[9] or '').strip()

    if raw_section and not raw_category:
        last_category = raw_section
        continue

    if raw_category:
        last_category = raw_category
    if raw_double_size:
        last_double_size = raw_double_size
    if raw_panel_size:
        last_panel_size = raw_panel_size

    category = last_category or raw_category
    double_size = last_double_size if category == '雙套管' else ''
    panel_size = raw_panel_size if raw_panel_size else (last_panel_size if category != '雙套管' else '')

    if part_name in ('', '台積料表'):
        continue

    entry = {
        'category': category,
        'doubleSize': double_size,
        'panelSize': panel_size,
        'material': raw_material,
        'partName': part_name,
        'lengthHint': length_hint,
        'quantityHint': qty_hint,
        'note': note,
    }

    if part_name == '無可對應料號':
        unmatched.append(entry)
    else:
        pipe_rules.append(entry)

OUTPUT.write_text(json.dumps({'rules': pipe_rules, 'unmatched': unmatched}, ensure_ascii=False, indent=2), encoding='utf-8')
print(f'Wrote {len(pipe_rules)} pipe rules, {len(unmatched)} unmatched entries -> {OUTPUT}')
