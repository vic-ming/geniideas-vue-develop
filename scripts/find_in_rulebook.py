from openpyxl import load_workbook


def norm(v):
    return "" if v is None else str(v)


def find(sheet, needle, max_hits=50):
    hits = []
    for r, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for c, v in enumerate(row, start=1):
            s = norm(v)
            if needle in s:
                hits.append((r, c, s[:160]))
                if len(hits) >= max_hits:
                    return hits
    return hits


def main():
    path = "src/assets/export/氣體別、聚賢料表、料號編碼規則 20251202.xlsx"
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in ["配件 (聚賢)", "其他元件 (聚賢)"]:
        ws = wb[sheet_name]
        for needle in ["TEE", "ELBOW", "DOUBLE TEE", "DOUBLE ELBOW", "REDUCER TEE", "GASKET", "NUT", "GLAND", "GAUGE", "Over Tube", "SPRING"]:
            hits = find(ws, needle, max_hits=8)
            if hits:
                print(f"{sheet_name} contains '{needle}':")
                for h in hits:
                    print("  ", h)
                print()


if __name__ == "__main__":
    main()




