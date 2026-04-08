import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path("src/assets/export")
RULE_BOOK = ROOT / "氣體別、聚賢料表、料號編碼規則 20251202.xlsx"
OUT = ROOT / "tsmc_fitting_rules.json"

wb = load_workbook(RULE_BOOK, read_only=True, data_only=True)
ws = wb["配件"]


def normalize(v):
    return "" if v is None else str(v).strip()


elbow = []
reducer_tee = []
reducer = []
cap = []
cross = []
tee = []
elbow45 = []

for row in ws.iter_rows(values_only=True):
    r = list(row)
    if len(r) <= 9:
        continue
    size = normalize(r[2])
    size2 = normalize(r[3])
    material = normalize(r[4])
    part = normalize(r[7])
    qty_rule = normalize(r[9])

    if not part or part in ("無可對應料號", "台積料表", "料表"):
        continue

    if part.startswith("ELBOW"):
        # size + material mapping, rule is length/3 ceil
        if size and material:
            elbow.append({
                "size": size,
                "material": material,
                "partName": part,
                "unit": "EA",
                "divisor": 3,
                "round": "ceil" if "進位" in qty_rule else "floor"
            })

    if part.startswith("REDUCER TEE"):
        if size and size2 and material:
            reducer_tee.append({
                "mainSize": size,
                "branchSize": size2,
                "material": material,
                "partName": part,
                "unit": "EA"
            })

    if part.startswith("REDUCER-"):
        if size and size2 and material:
            reducer.append({
                "fromSize": size,
                "toSize": size2,
                "material": material,
                "partName": part,
                "unit": "EA"
            })
    
    if part.startswith("CAP"):
        if size and material:
            cap.append({ "size": size, "material": material, "partName": part, "unit": "EA" })
            
    if part.startswith("CROSS"):
        if size and material:
            cross.append({ "size": size, "material": material, "partName": part, "unit": "EA" })
            
    if part.startswith("EQUAL TEE") or part.startswith("TEE 有分支"):
        if size and material:
            tee.append({ "size": size, "material": material, "partName": part, "unit": "EA" })
            
    if part.startswith("45 DEG"):
        if size and material:
            elbow45.append({ "size": size, "material": material, "partName": part, "unit": "EA" })

OUT.write_text(json.dumps({
    "elbow": elbow,
    "reducerTee": reducer_tee,
    "reducer": reducer,
    "cap": cap,
    "cross": cross,
    "tee": tee,
    "elbow45": elbow45
}, ensure_ascii=False, indent=2), encoding="utf-8")
