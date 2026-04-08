import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path("src/assets/export")
RULE_BOOK = ROOT / "氣體別、聚賢料表、料號編碼規則 20251202.xlsx"

OUT_VALVE = ROOT / "tsmc_valve_rules.json"
OUT_OTHER = ROOT / "tsmc_other_rules.json"

wb = load_workbook(RULE_BOOK, read_only=True, data_only=True)


def normalize(v):
    if v is None:
        return ""
    return str(v).strip()


def extract_valve_rules():
    ws = wb["閥件"]
    rules = []
    unmatched = []

    for row in ws.iter_rows(values_only=True):
        row = list(row)
        if len(row) < 9:
            continue
        enabled = normalize(row[2])
        connector = normalize(row[3])
        size = normalize(row[4])
        valve_type = normalize(row[5])
        part_name = normalize(row[7])
        qty = normalize(row[8])

        if enabled != "勾選":
            continue
        if not connector or not size or not valve_type:
            continue
        if not part_name or part_name in ("台積料表", "無可對應料號"):
            entry = {
                "connector": connector,
                "size": size,
                "valveType": valve_type,
                "partName": part_name,
                "qtyHint": qty,
            }
            if part_name == "無可對應料號":
                unmatched.append(entry)
            continue

        rules.append(
            {
                "connector": connector,
                "size": size,
                "valveType": valve_type,
                "partName": part_name,
                "unit": "EA",
            }
        )

    OUT_VALVE.write_text(
        json.dumps({"rules": rules, "unmatched": unmatched}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Valve rules: {len(rules)}, unmatched: {len(unmatched)} -> {OUT_VALVE}")


def extract_other_rules():
    ws = wb["其他元件"]

    stop_spacer = []
    spring = []
    over_tube = []
    gauge = []

    current = None

    for row in ws.iter_rows(values_only=True):
        row = list(row)
        if len(row) < 10:
            continue

        label = normalize(row[1])
        if label in ("STOP SPACER 收尾環", "SPRING 彈簧", "Over Tube 滑套", "GAUGE 壓力錶"):
            current = label
            continue

        if current == "STOP SPACER 收尾環":
            pipeline_type = normalize(row[2])
            double_size = normalize(row[3])
            part = normalize(row[7])
            if pipeline_type == "雙套管" and double_size and part and part != "無可對應料號":
                stop_spacer.append({"doubleSize": double_size, "partName": part, "unit": "EA"})

        elif current == "SPRING 彈簧":
            pipeline_type = normalize(row[2])
            double_size = normalize(row[3])
            part = normalize(row[7])
            if pipeline_type == "雙套管" and double_size and part and part != "無可對應料號":
                spring.append({"doubleSize": double_size, "partName": part, "unit": "EA"})

        elif current == "Over Tube 滑套":
            pipeline_type = normalize(row[2])
            double_size = normalize(row[3])
            part = normalize(row[7])
            if pipeline_type == "雙套管" and double_size and part and part != "無可對應料號":
                over_tube.append({"doubleSize": double_size, "partName": part, "unit": "EA"})

        elif current == "GAUGE 壓力錶":
            panel_size = normalize(row[3])
            panel_connector = normalize(row[4])
            material = normalize(row[5])
            part = normalize(row[7])
            if panel_size and panel_connector and material and part and part != "無可對應料號":
                gauge.append(
                    {
                        "panelSize": panel_size,
                        "panelConnector": panel_connector,
                        "material": material,
                        "partName": part,
                        "unit": "EA",
                    }
                )

    OUT_OTHER.write_text(
        json.dumps(
            {
                "stopSpacer": stop_spacer,
                "spring": spring,
                "overTube": over_tube,
                "gauge": gauge,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(
        f"Other rules: stopSpacer={len(stop_spacer)}, spring={len(spring)}, overTube={len(over_tube)}, gauge={len(gauge)} -> {OUT_OTHER}"
    )


extract_valve_rules()
extract_other_rules()
