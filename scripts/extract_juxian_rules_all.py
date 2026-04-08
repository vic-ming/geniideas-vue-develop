import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path("src/assets/export")
RULE_BOOK = ROOT / "氣體別、聚賢料表、料號編碼規則 20251202.xlsx"

OUT_PIPE = ROOT / "juxian_pipe_rules.json"
OUT_VALVE = ROOT / "juxian_valve_rules.json"
OUT_OTHER = ROOT / "juxian_other_rules.json"
OUT_FITTING = ROOT / "juxian_fitting_rules.json"


def norm(v):
    return "" if v is None else str(v).strip()


def guess_brand_from_name(name: str) -> str:
    """
    Most 聚賢 names follow: ..._<BRAND>_無
    So we take the token before the last underscore token.
    """
    s = norm(name)
    if not s:
        return ""
    parts = s.split("_")
    if len(parts) < 2:
        return ""
    # last token is usually '無'
    return parts[-2].strip()


def scan_adjacent_pairs(row_values, start_idx: int):
    """
    Yield (name, partNo) for any adjacent non-empty cells starting at start_idx.
    """
    r = [norm(x) for x in row_values]
    for i in range(start_idx, len(r) - 1):
        name = r[i]
        part_no = r[i + 1]
        if not name or not part_no:
            continue
        if "氣體材料" not in name and "G" not in part_no:
            # skip obvious non rule pairs (still keep if looks like a part number)
            continue
        yield name, part_no


def write_json(path: Path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def extract_pipe(wb):
    ws = wb["管線 (聚賢)"]
    rules = []
    unmatched = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = list(row)
        if len(r) < 7:
            continue
        pipeline_type = norm(r[2])
        size = norm(r[4])
        material = norm(r[6])
        if not pipeline_type or not size or not material:
            continue

        found_any = False
        for name, part_no in scan_adjacent_pairs(r, start_idx=7):
            found_any = True
            brand = guess_brand_from_name(name)
            entry = {
                "pipelineType": pipeline_type,
                "size": size,
                "material": material,
                "brand": brand,
                "name": name,
                "partNo": part_no,
                "unit": "M",
            }
            if name == "無可對應料號" or part_no == "無可對應料號":
                unmatched.append(entry)
            else:
                rules.append(entry)

        if not found_any:
            # still record as unmatched if template marks it
            pass

    return {"rules": rules, "unmatched": unmatched}


def extract_valve(wb):
    ws = wb["閥件 (聚賢)"]
    rules = []
    unmatched = []

    # header row defines brand pairs starting from first repeated brand cell
    header = [norm(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    brand_cols = []  # (brand, name_col_idx, part_col_idx)
    for idx in range(0, len(header) - 1):
        b1 = header[idx]
        b2 = header[idx + 1]
        if b1 and b1 == b2:
            brand_cols.append((b1, idx, idx + 1))

    for row in ws.iter_rows(min_row=2, values_only=True):
        r = list(row)
        if len(r) < 6:
            continue
        checked = norm(r[2])
        connector = norm(r[3])
        size = norm(r[4])
        raw_valve_type = norm(r[5])
        if checked != "勾選" or not connector or not size:
            continue

        for brand, n_idx, p_idx in brand_cols:
            name = norm(r[n_idx]) if n_idx < len(r) else ""
            part_no = norm(r[p_idx]) if p_idx < len(r) else ""
            if not name and not part_no:
                continue

            # regulator rows in the rulebook use '勾選' in valveType column; infer from name
            valve_type = raw_valve_type
            if "REGULATOR" in name.upper():
                valve_type = "REGULATOR"

            # skip rows without a meaningful valve type (except inferred regulator above)
            if not valve_type or valve_type == "勾選":
                continue

            entry = {
                "connector": connector,
                "size": size,
                "valveType": valve_type,
                "brand": brand,
                "name": name,
                "partNo": part_no,
                "unit": "EA",
            }
            if (
                name == "無可對應料號"
                or part_no == "無可對應料號"
                or not name
                or not part_no
            ):
                unmatched.append(entry)
            else:
                rules.append(entry)

    return {"rules": rules, "unmatched": unmatched}


def extract_other(wb):
    ws = wb["其他元件 (聚賢)"]
    sections = {
        "stopSpacer": [],
        "spring": [],
        "overTube": [],
        "gauge": [],
        "gland": [],
        "nut": [],
        "gasket": [],
    }
    unmatched = []

    current = "stopSpacer"
    for row in ws.iter_rows(min_row=1, values_only=True):
        r = [norm(x) for x in row]
        if not any(r):
            continue
        title = r[1] if len(r) > 1 else ""
        if title.startswith("STOP SPACER"):
            current = "stopSpacer"
            continue
        if title.startswith("SPRING"):
            current = "spring"
            continue
        if title.startswith("Over Tube"):
            current = "overTube"
            continue
        if title.startswith("GAUGE"):
            current = "gauge"
            continue
        if title == "GLAND":
            current = "gland"
            continue
        if title == "NUT":
            current = "nut"
            continue
        if title == "GASKET":
            current = "gasket"
            continue

        # section headers often have A2/A6 labels in col3/4 - skip
        if len(r) > 2 and r[2].startswith("A2"):
            continue

        # parse conditions by section
        base = {}
        if current in ("stopSpacer", "spring", "overTube"):
            pipeline_type = r[2] if len(r) > 2 else ""
            double_size = r[3] if len(r) > 3 else ""
            if not pipeline_type or not double_size:
                continue
            base = {"pipelineType": pipeline_type, "doubleSize": double_size}
            if current == "stopSpacer":
                branch_flag = r[5] if len(r) > 5 else ""
                base["hasBranchPanel"] = branch_flag == "勾選"
        elif current in ("gauge", "gland", "nut", "gasket"):
            panel_size = r[3] if len(r) > 3 else ""
            panel_connector = r[4] if len(r) > 4 else ""
            material = r[5] if len(r) > 5 else ""
            if not panel_size or not panel_connector or not material:
                continue
            base = {
                "panelSize": panel_size,
                "panelConnector": panel_connector,
                "material": material,
            }
        else:
            continue

        any_pair = False
        for name, part_no in scan_adjacent_pairs(r, start_idx=7):
            any_pair = True
            brand = guess_brand_from_name(name)
            entry = {
                **base,
                "brand": brand,
                "name": name,
                "partNo": part_no,
                "unit": "EA",
            }
            if name == "無可對應料號" or part_no == "無可對應料號":
                unmatched.append({"section": current, **entry})
            else:
                sections[current].append(entry)
        if not any_pair:
            # some rows might be blank for many brands; ignore
            pass

    return {"sections": sections, "unmatched": unmatched}


def extract_fitting(wb):
    ws = wb["配件 (聚賢)"]
    sections = {
        "reducer": [],
        "reducerTee": [],
        "tee": [],
        "elbow": [],
        "doubleTee": [],
        "doubleElbow": [],
    }
    unmatched = []

    current = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        r = [norm(x) for x in row]
        if not any(r):
            continue
        title = r[1] if len(r) > 1 else ""
        if title.startswith("REDUCER 無分支"):
            current = "reducer"
            continue
        if title.startswith("REDUCER TEE"):
            current = "reducerTee"
            continue
        if title.startswith("TEE 有分支"):
            current = "tee"
            continue
        if title.startswith("ELBOW"):
            current = "elbow"
            continue
        if title.startswith("DOUBLE TEE"):
            current = "doubleTee"
            continue
        if title.startswith("DOUBLE ELBOW"):
            current = "doubleElbow"
            continue

        # skip header rows that define condition labels
        if len(r) > 2 and (r[2].startswith("A") or r[2].startswith("B") or r[2].startswith("C") or r[2].startswith("選單")):
            continue

        base = {}
        if current in ("reducer", "reducerTee"):
            from_size = r[2] if len(r) > 2 else ""
            to_size = r[3] if len(r) > 3 else ""
            material = r[4] if len(r) > 4 else ""
            if not from_size or not to_size or not material:
                continue
            base = {"fromSize": from_size, "toSize": to_size, "material": material}
        elif current == "tee":
            # TEE 有分支：前尺寸=後尺寸，使用同一尺寸 + 材質
            size = r[3] if len(r) > 3 else ""
            material = r[4] if len(r) > 4 else ""
            if not size or not material:
                continue
            base = {"size": size, "material": material}
        elif current == "elbow":
            size = r[2] if len(r) > 2 else ""
            material = r[4] if len(r) > 4 else ""
            if not size or not material:
                continue
            base = {"size": size, "material": material}
        elif current in ("doubleTee", "doubleElbow"):
            pipeline_type = r[2] if len(r) > 2 else ""
            double_size = r[3] if len(r) > 3 else ""
            if not pipeline_type or not double_size:
                continue
            base = {"pipelineType": pipeline_type, "doubleSize": double_size}
        else:
            continue

        any_pair = False
        for name, part_no in scan_adjacent_pairs(r, start_idx=6):
            any_pair = True
            brand = guess_brand_from_name(name)
            entry = {
                **base,
                "brand": brand,
                "name": name,
                "partNo": part_no,
                "unit": "EA",
            }
            if name == "無可對應料號" or part_no == "無可對應料號":
                unmatched.append({"section": current, **entry})
            else:
                sections[current].append(entry)
        if not any_pair:
            pass

    return {"sections": sections, "unmatched": unmatched}


def main():
    wb = load_workbook(RULE_BOOK, read_only=True, data_only=True)
    pipe = extract_pipe(wb)
    valve = extract_valve(wb)
    other = extract_other(wb)
    fitting = extract_fitting(wb)

    write_json(OUT_PIPE, pipe)
    write_json(OUT_VALVE, valve)
    write_json(OUT_OTHER, other)
    write_json(OUT_FITTING, fitting)

    print(f"Wrote pipe: rules={len(pipe['rules'])} unmatched={len(pipe['unmatched'])} -> {OUT_PIPE}")
    print(f"Wrote valve: rules={len(valve['rules'])} unmatched={len(valve['unmatched'])} -> {OUT_VALVE}")
    for k, v in other["sections"].items():
        print(f"Wrote other {k}: {len(v)}")
    print(f"Wrote other unmatched={len(other['unmatched'])} -> {OUT_OTHER}")
    for k, v in fitting["sections"].items():
        print(f"Wrote fitting {k}: {len(v)}")
    print(f"Wrote fitting unmatched={len(fitting['unmatched'])} -> {OUT_FITTING}")


if __name__ == "__main__":
    main()

